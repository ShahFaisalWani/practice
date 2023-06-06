import sys
args = sys.argv[1:]
file_path = args[0]

col_dictionary = {
     "ISBN" : ["บาร์โค้ด","ISBN","รหัสสินค้า"],
     "title": ["ชื่อ","เรื่อง", "ชื่อหนังสือ"],
     "desc": ["เนื้อหา","คำโปรย"],
     "author": ["ผู้แต่ง"],
     "translator": ["ผู้แปล"],
     "publisher": ["สำนักพิมพ์","สนพ"],
     "genre": ["หมวด","ประเภท"],
     "price": ["ราคาปก"],
     "weight": ["weight","น้ำหนัก"] 
}

target_columns = list(col_dictionary.values())

import pandas as pd
import openpyxl
from fuzzywuzzy import fuzz, process
import json

# Load the Excel file
workbook = openpyxl.load_workbook(file_path)

# Process each worksheet
worksheet = workbook.worksheets[0]

# Find the start row and column of the table
start_row = None
start_col = None

for row in worksheet.iter_rows():
    if row[0].value is not None and row[1].value is not None:
        start_row = row[2].row
        start_col = row[0].column
        break

# Find the end row and column of the table
end_row = start_row
end_col = start_col

for row in worksheet.iter_rows(min_row=start_row, min_col=start_col):
    row_values = [cell.value for cell in row]
    if all(value is None for value in row_values):
        break
    end_row = row[0].row
    end_col = max(cell.column for cell in row if cell.value is not None)

# Get the range of the table
table_range = (
    start_row, start_col,
    end_row, end_col
)

# Read the table content into a DataFrame
data = list(worksheet.iter_rows(
    min_row=table_range[0], min_col=table_range[1],
    max_row=table_range[2], max_col=table_range[3],
    values_only=True  # Retrieve cell values instead of cell objects
))

# Create a DataFrame from the table data
df = pd.DataFrame(data[1:], columns=data[0])

def getKeyIndex(val):
    return next((key for key, value in col_dictionary.items() if isinstance(value, list) and val in value), None)

matches = {}
def find_best_match(string, choices, threshold=65):
    for word in string:
        best_match, score = process.extractOne(word, choices, scorer=fuzz.token_sort_ratio)
        if score >= threshold:               
            if(getKeyIndex(word) not in matches):
                matches[getKeyIndex(word)] = {"string":word,"target":best_match,"score":score}
            elif(matches[getKeyIndex(word)]['score'] < score):
                matches[getKeyIndex(word)] = {"string":word,"target":best_match,"score":score}

for column in target_columns:
    find_best_match(column, df.columns)


db_columns = list(col_dictionary.keys())

renameCol = {}
newmatch = []
for i in range(len(matches)):  
    val = list(matches.values())[i]
    renameCol[val['target']] = getKeyIndex(val['string'])
    newmatch.append(val['target'])

df = df[newmatch]
df.rename(columns=renameCol, inplace = True)


for i in range(len(df['ISBN'])):
    if(len(str(df['ISBN'][i]))>13):
        df['ISBN'][i] = ''.join(str(df['ISBN'][i]).split('-')).strip()

if 'weight' in df.columns:
    if(df['weight'][0]>10):
        df['weight'] = df['weight'] / 1000


print(json.dumps({"matches": df.to_json(orient='records')}))

# Save the extracted DataFrame as an Excel file
df.to_excel("output.xlsx", index=False)







